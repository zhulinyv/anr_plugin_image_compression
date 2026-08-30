import shutil
from pathlib import Path

import cv2
import numpy as np
import openpyxl
import ujson as json
from openpyxl.drawing.image import Image as OPXIMG
from openpyxl.styles import Alignment
from PIL import Image as PILIMG

from utils.image_tools import get_image_information, revert_image_info
from utils.logger import logger


def get_image_paths(directory: str | Path) -> list[Path]:
    valid_exts = {".png", ".jpg", ".jpeg", ".webp"}
    return [p for p in Path(directory).iterdir() if p.is_file() and p.suffix.lower() in valid_exts]


def resolve_images(path: str | Path | list) -> list[Path]:
    """目录 -> 目录内全部图片; 单张图片文件 -> 该文件; 传入列表时先处理文件再处理目录。"""
    valid_exts = {".png", ".jpg", ".jpeg", ".webp"}
    paths = path if isinstance(path, (list, tuple)) else [path]
    images: list[Path] = []
    for item in paths:
        if not item:
            continue
        p = Path(item)
        if p.is_file():
            if p.suffix.lower() in valid_exts:
                images.append(p)
        elif p.is_dir():
            images.extend(get_image_paths(p))
    # 去重 (保留顺序: 先图片, 再目录)
    seen = set()
    result: list[Path] = []
    for img in images:
        key = str(img.resolve()) if img.exists() else str(img)
        if key not in seen:
            seen.add(key)
            result.append(img)
    return result


def return_pnginfo(image: PILIMG.Image) -> dict:
    pnginfo = get_image_information(image)
    comment = json.loads(pnginfo.get("Comment", "{}"))

    return {
        "prompt": comment.get("prompt", ""),
        "uc": comment.get("uc", ""),
        "width": str(comment.get("width", "")),
        "height": str(comment.get("height", "")),
        "steps": comment.get("steps", ""),
        "scale": comment.get("scale", ""),
        "noise_schedule": comment.get("noise_schedule", ""),
        "sampler": comment.get("sampler", ""),
        "sm": comment.get("sm", ""),
        "sm_dyn": comment.get("sm_dyn", ""),
        "skip_cfg_above_sigma": bool(comment.get("skip_cfg_above_sigma", False)),
        "dynamic_thresholding": comment.get("dynamic_thresholding", ""),
        "seed": comment.get("seed", ""),
    }


def _image_compression(format_: str, image_path: Path) -> Path:
    cv2_image = cv2.imread(str(image_path), cv2.IMREAD_UNCHANGED)

    if format_ == "jpg":
        # 利用 numpy 广播机制高效处理 Alpha 通道 (透明背景转纯白)
        if cv2_image.shape[2] == 4:
            alpha = cv2_image[:, :, 3:] / 255.0  # 保持三维以便广播
            bgr = cv2_image[:, :, :3]
            cv2_image = (bgr * alpha + 255 * (1 - alpha)).astype(np.uint8)

        compression_params = [cv2.IMWRITE_JPEG_QUALITY, 75]
        output_dir = Path("./outputs")
        output_dir.mkdir(exist_ok=True)

        temp_path = output_dir / f"temp_compression.{format_}"
        cv2.imwrite(str(temp_path), cv2_image, compression_params)
        new_path = image_path.with_name(f"{image_path.stem}_compression.{format_}")
        shutil.move(str(temp_path), str(new_path))
    elif format_ == "png":
        # compression_params = [cv2.IMWRITE_PNG_COMPRESSION, 9]
        with PILIMG.open(image_path) as img:
            if img.mode == 'RGBA':
                alpha = img.getchannel('A')
                img_p = img.convert('P', palette=PILIMG.ADAPTIVE, colors=256)
            else:
                img_p = img.convert('P', palette=PILIMG.ADAPTIVE, colors=256)

            new_path = image_path.with_name(f"{image_path.stem}_compression.png")

            img_p.save(
                new_path,
                format="PNG",
                optimize=True,
                compress_level=9
            )
            revert_image_info(str(image_path), str(new_path))
    elif format_ == "webp":
        with PILIMG.open(image_path) as img:
            if img.mode != "RGBA":
                img = img.convert("RGBA")

            new_path = image_path.with_name(f"{image_path.stem}_compression.webp")

            img.save(
                new_path,
                "WEBP",
                quality=80,
                lossless=False,
                method=6
            )
            revert_image_info(str(image_path), str(new_path))

    else:
        raise ValueError(f"Unsupported format: {format_}")

    return new_path


def image_compression(format_: str, image_path: str | Path) -> dict:
    """压缩图片, 返回 {"message": ..., "images": [新图片路径...]} 供前端直接展示。"""
    image_list = resolve_images(image_path)
    logger.info(f"开始压缩, 共 {len(image_list)} 张图片...")
    compressed: list[Path] = []
    for image in image_list:
        logger.info(f"正在压缩 {image.name} ...")
        new_path = _image_compression(format_, image)
        if new_path and new_path.exists():
            compressed.append(new_path)

    logger.success(f"压缩完成! 共 {len(compressed)} 张")
    # as_posix: 统一用正斜杠, 前端按 "/" 切分文件名 / 打开目录更稳妥
    return {"message": f"压缩完成! 共 {len(compressed)} 张", "images": [p.as_posix() for p in compressed]}


def image_organization(format_: str, image_path: str | Path, switch: bool) -> dict:
    """整理图片为 excel 表格, 返回 {"message": ..., "dir": 保存目录} 供前端打开目录。"""
    logger.info("正在整理图片...")
    image_list = resolve_images(image_path)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        "图片",
        "正面提示词",
        "负面提示词",
        "分辨率",
        "采样步数",
        "提示词相关性",
        "调度器",
        "采样器",
        "sm",
        "sm_dyn",
        "variety",
        "decrisp",
        "随机种子",
    ]
    ws.append(headers)
    ws.row_dimensions[1].height = 50

    for col_idx, width in enumerate([40, 40, 40] + [20] * 10, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = (
            width
        )

    alignment = Alignment(horizontal="center", vertical="center", wrapText=True)

    for row_idx, image in enumerate(image_list, start=2):
        ws.row_dimensions[row_idx].height = 300

        with PILIMG.open(image) as pilimg:
            info = return_pnginfo(pilimg)
            w, h = pilimg.size

        insert_image_path = image
        if switch:
            insert_image_path = _image_compression(format_, image)

        opx_img = OPXIMG(str(insert_image_path))
        opx_img.width, opx_img.height = 260, int(260 / w * h)
        ws.add_image(opx_img, f"A{row_idx}")

        w = info.get("width", "")
        h = info.get("height", "")

        row_data = [
            None,  # 图片占位 A 列
            info.get("prompt", ""),
            info.get("uc", ""),
            f"{w}x{h}",
            info.get("steps", ""),
            info.get("scale", ""),
            info.get("noise_schedule", ""),
            info.get("sampler", ""),
            info.get("sm", ""),
            info.get("sm_dyn", ""),
            info.get("skip_cfg_above_sigma", ""),  # 原代码中的 variety 对应位置
            info.get("dynamic_thresholding", ""),  # 原代码中的 decrisp 对应位置
            info.get("seed", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = alignment

    for cell in ws[1]:
        cell.alignment = alignment

    # 保存目录: 提供目录时保存到该目录, 否则保存到第一张图片所在目录
    paths = image_path if isinstance(image_path, (list, tuple)) else [image_path]
    save_dir = next((Path(p) for p in paths if p and Path(p).is_dir()), None)
    if save_dir is None:
        save_dir = image_list[0].parent if image_list else Path("./outputs")
    save_path = save_dir / "organization.xlsx"
    wb.save(save_path)
    logger.success(f"整理完成! 表格已保存到 {save_path}")
    return {"message": "整理完成!", "dir": save_dir.as_posix()}
